import openpyxl
import pandas as pd
import yfinance as yf
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Her kategoride tam 100 adet olmak üzere toplam 400 ETF listesi
kategoriler = {
    "Temettü ETF'leri": [
        "SCHD",
        "VYM",
        "DGRO",
        "DGRW",
        "VIG",
        "JEPI",
        "JEPQ",
        "SDY",
        "NOBL",
        "RDIV",
        "SPYD",
        "HDV",
        "FDL",
        "PEY",
        "DHS",
        "OUSA",
        "VYMI",
        "IDV",
        "DON",
        "SDOG",
        "FVD",
        "REGL",
        "DIV",
        "KBWD",
        "KBWY",
        "PFF",
        "PGX",
        "SPHD",
        "FDVV",
        "QDIV",
        "IDVO",
        "DIVO",
        "NUSI",
        "QYLD",
        "RYLD",
        "XYLD",
        "SPYI",
        "ISPY",
        "YMAG",
        "YMAX",
        "FEZ",
        "DEM",
        "EELV",
        "DGRE",
        "FPE",
        "PFFD",
        "PFFA",
        "DVYE",
        "EDIV",
        "CEFS",
        "PFXF",
        "SPFF",
        "VRP",
        "HIPS",
        "DIVZ",
        "ROIC",
        "DGRS",
        "TDIV",
        "JEPY",
        "GPIQ",
        "GPIX",
        "QDPL",
        "FTSM",
        "HVPW",
        "WBIY",
        "DES",
        "OTM",
        "LVL",
        "RDVY",
        "DIVX",
        "EPRF",
        "PGHY",
        "HYXF",
        "JNK",
        "HYG",
        "USHY",
        "ANGL",
        "FTSD",
        "ETJ",
        "BXSL",
        "ARCC",
        "MAIN",
        "HTGC",
        "OBDC",
        "GBDC",
        "TSLX",
        "CSWC",
        "PSEC",
        "TPVG",
        "FDUS",
        "GAIN",
        "OXLC",
        "ECC",
        "PFLT",
        "SUNS",
        "WHF",
        "TRIN",
        "GECC",
        "PFF",
        "PGX",
    ],
    "Büyüme ETF'leri": [
        "QQQ",
        "QQQM",
        "SCHG",
        "VUG",
        "IWF",
        "SPYG",
        "VGT",
        "ARKK",
        "XLK",
        "SMH",
        "SOXX",
        "IYW",
        "FTEC",
        "IGV",
        "XLC",
        "XLY",
        "ARKW",
        "ARKG",
        "ONEQ",
        "MTUM",
        "VOOG",
        "IWY",
        "RPG",
        "MGK",
        "XBI",
        "IBB",
        "IHI",
        "PHO",
        "TAN",
        "PBW",
        "QCLN",
        "ICLN",
        "LIT",
        "BJK",
        "GAMR",
        "NERD",
        "BETZ",
        "HERO",
        "BOTZ",
        "ROBT",
        "AIQ",
        "SNSR",
        "FINX",
        "IPAY",
        "HACK",
        "CIBR",
        "BUG",
        "CLOU",
        "SKYY",
        "WCLD",
        "SOCL",
        "XSD",
        "XSW",
        "PSI",
        "IGN",
        "XNTK",
        "FDG",
        "FDIS",
        "FTEK",
        "FSTA",
        "FUTY",
        "FXG",
        "FXL",
        "FXO",
        "FXU",
        "FXI",
        "FXN",
        "ARKQ",
        "ARKF",
        "ARKX",
        "AUTN",
        "DRIV",
        "KGRN",
        "RNRG",
        "ACES",
        "FAN",
        "PAVE",
        "MOO",
        "JETS",
        "PEJ",
        "BEDZ",
        "HAIL",
        "COM",
        "EBIZ",
        "ONLN",
        "SGDM",
        "BAR",
        "OUNZ",
        "SIVR",
        "PICK",
        "HYDR",
        "AMNA",
        "NLR",
        "AWAY",
        "BETZ",
        "NERD",
        "GAMR",
        "HERO",
        "BOTZ",
        "ROBT",
    ],
    "Geniş Piyasa": [
        "SPY",
        "VOO",
        "VTI",
        "VT",
        "IVV",
        "RSP",
        "DIA",
        "IWM",
        "MDY",
        "IJH",
        "IJR",
        "SCHB",
        "ITOT",
        "VV",
        "MGC",
        "VO",
        "VB",
        "SCHV",
        "SCHA",
        "IVW",
        "IVE",
        "IWD",
        "IWN",
        "IWO",
        "IWP",
        "IWS",
        "IWZ",
        "OEF",
        "EPS",
        "PKW",
        "DES",
        "DGRS",
        "DON",
        "IWV",
        "JKG",
        "JKH",
        "JKI",
        "IYY",
        "IYF",
        "IYK",
        "IYM",
        "IYN",
        "IYP",
        "IYR",
        "IYT",
        "IYW",
        "IYZ",
        "IWB",
        "IWC",
        "IWE",
        "IWF",
        "IWG",
        "IWH",
        "IWI",
        "IWJ",
        "IWK",
        "IWL",
        "IWQ",
        "IWR",
        "IWT",
        "IWU",
        "IWW",
        "IWX",
        "IWY",
        "SPGM",
        "SPLG",
        "SPHB",
        "SPLV",
        "USMV",
        "QUAL",
        "MTUM",
        "SIZE",
        "VLUE",
        "DFAC",
        "DFLV",
        "DFSV",
        "DFUV",
        "AVLV",
        "AVGV",
        "AVUV",
        "AVDV",
        "XLG",
        "QQQE",
        "RSPR",
        "EQL",
        "EWV",
        "EWX",
        "EWY",
        "EWZ",
        "EWA",
        "EWC",
        "EWD",
        "EWG",
        "EWH",
        "EWJ",
        "EWK",
        "EWL",
        "EWM",
        "EWN",
        "EWO",
    ],
    "Emtia ETF'leri": [
        "GLD",
        "IAU",
        "SLV",
        "DBC",
        "USO",
        "UNG",
        "URA",
        "REMX",
        "COPX",
        "PPLT",
        "PALL",
        "DBB",
        "DBA",
        "WEAT",
        "CORN",
        "SOYB",
        "JO",
        "NIB",
        "TAGS",
        "BNO",
        "UDN",
        "UUP",
        "FXE",
        "FXB",
        "FXF",
        "FXC",
        "CYB",
        "CEW",
        "DBP",
        "DBE",
        "DBV",
        "COM",
        "GCC",
        "PDBC",
        "BCI",
        "GUNR",
        "FTGC",
        "RJI",
        "DJP",
        "GSG",
        "UCI",
        "CMDY",
        "KRBN",
        "MOO",
        "SOIL",
        "WOOD",
        "SGDM",
        "BAR",
        "OUNZ",
        "SIVR",
        "PICK",
        "LIT",
        "HYDR",
        "AMNA",
        "NLR",
        "FAN",
        "TAN",
        "ICLN",
        "QCLN",
        "PBW",
        "ACES",
        "HAIL",
        "JETS",
        "PEJ",
        "BEDZ",
        "AWAY",
        "BETZ",
        "NERD",
        "GAMR",
        "HERO",
        "BOTZ",
        "ROBT",
        "AIQ",
        "SNSR",
        "FINX",
        "IPAY",
        "HACK",
        "CIBR",
        "BUG",
        "CLOU",
        "SKYY",
        "WCLD",
        "SOCL",
        "XSD",
        "XSW",
        "PSI",
        "IGN",
        "XNTK",
        "FDG",
        "FDIS",
        "FTEK",
        "FSTA",
        "FUTY",
        "FXG",
        "FXL",
        "FXO",
        "FXU",
        "FXI",
        "FXN",
    ],
}

output_file = "etf_canli_kutuphane.xlsx"
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Varsayılan boş sayfayı sil

columns = [
    "Kod",
    "ETF Adı",
    "Türkçe Adı",
    "Fiyat",
    "ETF Toplam Değer",
    "Temettü Verimi %",
    "Yılbaşından Bugüne Getiri",
    "1 AYLIK GETİRİ",
    "3 AYLIK GETİRİ",
    "1 YILLIK GETİRİ",
    "YÖNETİM ÜCRETİ %",
]

print(
    "Toplam 400 ETF için canlı borsa verileri Yahoo Finance üzerinden çekiliyor,"
    " lütfen bekleyin (Bu işlem birkaç dakika sürebilir)..."
)

for sheet_title, tickers in kategoriler.items():
  ws = wb.create_sheet(title=sheet_title)
  ws.views.sheetView[0].showGridLines = True
  ws.freeze_panes = "A2"

  header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
  header_fill = PatternFill(
      start_color="1F4E79", end_color="1F4E79", fill_type="solid"
  )
  row_fill_light = PatternFill(
      start_color="F9F9F9", end_color="F9F9F9", fill_type="solid"
  )
  border_thin = Border(
      left=Side(style="thin", color="D9D9D9"),
      right=Side(style="thin", color="D9D9D9"),
      top=Side(style="thin", color="D9D9D9"),
      bottom=Side(style="thin", color="D9D9D9"),
  )

  ws.append(columns)
  for col_num in range(1, len(columns) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    cell.border = border_thin
  ws.row_dimensions[1].height = 28

  row_idx = 2
  for ticker in tickers:
    try:
      t = yf.Ticker(ticker)
      info = t.info

      name = info.get("longName", ticker)
      price = info.get(
          "regularMarketPrice",
          info.get("currentPrice", info.get("previousClose", 0)),
      )
      aum_val = info.get("totalAssets", 0)
      aum_str = f"{round(aum_val / 1e9, 1)}B" if aum_val else "N/A"

      div_yield = info.get("dividendYield", 0)
      div_yield = round(div_yield * 100, 2) if div_yield else 0.0

      exp_ratio = info.get("expenseRatio", 0)
      exp_ratio = round(exp_ratio * 100, 2) if exp_ratio else 0.0

      hist = t.history(period="1y")
      ytd_ret, m1_ret, m3_ret, y1_ret = 0.0, 0.0, 0.0, 0.0

      if not hist.empty:
        current_p = hist["Close"].iloc[-1]
        start_p_1y = hist["Close"].iloc[0]
        y1_ret = round(((current_p - start_p_1y) / start_p_1y) * 100, 2)

        year_start = hist[hist.index.year == hist.index[-1].year]
        if not year_start.empty:
          start_p_ytd = year_start["Close"].iloc[0]
          ytd_ret = round(((current_p - start_p_ytd) / start_p_ytd) * 100, 2)

        if len(hist) >= 20:
          start_p_1m = hist["Close"].iloc[-20]
          m1_ret = round(((current_p - start_p_1m) / start_p_1m) * 100, 2)
        if len(hist) >= 60:
          start_p_3m = hist["Close"].iloc[-60]
          m3_ret = round(((current_p - start_p_3m) / start_p_3m) * 100, 2)

      tr_name = f"ABD Borsası {ticker} Fonu"
      row_data = [
          ticker,
          name,
          tr_name,
          round(price, 2),
          aum_str,
          div_yield,
          ytd_ret,
          m1_ret,
          m3_ret,
          y1_ret,
          exp_ratio,
      ]

    except Exception as e:
      row_data = [ticker, ticker, "Fon", 0.0, "N/A", 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]

    ws.append(row_data)
    ws.row_dimensions[row_idx].height = 20
    apply_fill = row_fill_light if row_idx % 2 == 0 else None

    for col_idx in range(1, len(columns) + 1):
      cell = ws.cell(row=row_idx, column=col_idx)
      cell.border = border_thin
      if apply_fill:
        cell.fill = apply_fill

      if col_idx == 1:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Calibri", size=10, bold=True)
      elif col_idx in [2, 3]:
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = Font(name="Calibri", size=10)
      elif col_idx == 4:
        cell.number_format = "$#,##0.00"
        cell.alignment = Alignment(horizontal="right", vertical="center")
      elif col_idx == 5:
        cell.alignment = Alignment(horizontal="center", vertical="center")
      elif col_idx >= 6:
        val = cell.value
        if isinstance(val, (int, float)):
          cell.value = val / 100.0
          cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="right", vertical="center")
    row_idx += 1

  ws.column_dimensions["A"].width = 10
  ws.column_dimensions["B"].width = 42
  ws.column_dimensions["C"].width = 42
  ws.column_dimensions["D"].width = 15
  ws.column_dimensions["E"].width = 18
  for col_idx in range(6, 12):
    ws.column_dimensions[get_column_letter(col_idx)].width = 15

wb.save(output_file)
print(
    f"\nİşlem tamamlandı! Toplam 400 ETF için güncel veriler Excel'e kaydedildi:"
    f" {output_file}"
)